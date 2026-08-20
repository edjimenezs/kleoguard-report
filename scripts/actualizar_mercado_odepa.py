#!/usr/bin/env python3
"""Actualiza las series públicas de novillos y vaquillas.

La fuente es el boletín semanal público de ODEPA/AFECH. Este programa nunca
consulta ni escribe información de productores, correos, SII o SAG.

Al ejecutarse de forma diaria conserva la historia ya publicada y añade sólo
jornadas oficiales nuevas. Si no cambió ningún dato comparable, no reescribe el
JSON: así GitHub Actions no genera commits falsos sólo por cambiar la hora.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from statistics import mean
from urllib.parse import urljoin
from urllib.request import Request, urlopen

from openpyxl import load_workbook


BOLETIN_URL = (
    "https://www.odepa.gob.cl/contenidos-rubro/boletines-del-rubro/"
    "boletin-semanal-de-precios-asoc-gremial-de-ferias-ganaderas"
)
DESTINO = Path(__file__).resolve().parents[1] / "market-data.json"
MAX_BOLETINES_RECIENTES = 10
SHEET_NAME = "Promedio (5 primeros precios)"
CATEGORIAS = (
    "Novillo Gordo",
    "Novillo Engorda",
    "Vaquilla Gorda",
    "Vaquilla Engorda",
)
SERIE_DESDE = date(2022, 1, 1)


def texto(valor: object) -> str:
    return re.sub(r"\s+", " ", str(valor or "")).strip()


def descargar(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": "KleoGuard-public-market/1.1"})
    with urlopen(request, timeout=45) as response:  # nosec B310: URL pública y fija.
        return response.read()


def links_xlsx(historial_completo: bool) -> list[str]:
    pagina = descargar(BOLETIN_URL).decode("utf-8", errors="ignore")
    hallados = re.findall(r'''href=["']([^"']+\.xlsx(?:\?[^"']*)?)["']''', pagina, flags=re.I)
    unicos: list[str] = []
    for enlace in hallados:
        enlace = urljoin(BOLETIN_URL, enlace.replace("&amp;", "&"))
        if enlace not in unicos:
            unicos.append(enlace)
    if not unicos:
        raise RuntimeError("ODEPA no entregó enlaces XLSX en la página del boletín.")
    return unicos if historial_completo else unicos[:MAX_BOLETINES_RECIENTES]


def numero(valor: object) -> float | None:
    if isinstance(valor, (int, float)):
        return float(valor) if float(valor) > 0 else None
    try:
        convertido = float(str(valor).replace(".", "").replace(",", "."))
    except (TypeError, ValueError):
        return None
    return convertido if convertido > 0 else None


def leer_boletin(contenido: bytes, url: str) -> list[dict[str, object]]:
    libro = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    if SHEET_NAME not in libro.sheetnames:
        raise RuntimeError(f"El XLSX cambió: no existe la hoja {SHEET_NAME!r}.")
    hoja = libro[SHEET_NAME]

    encabezado_fila = None
    encabezados: list[str] = []
    for indice, fila in enumerate(hoja.iter_rows(max_row=20, values_only=True), start=1):
        candidatos = [texto(celda) for celda in fila]
        if "Feria" in candidatos and "Fecha" in candidatos:
            encabezado_fila, encabezados = indice, candidatos
            break
    if encabezado_fila is None:
        raise RuntimeError("No se encontró el encabezado Feria/Fecha en el XLSX.")

    columnas = {nombre: posicion for posicion, nombre in enumerate(encabezados)}
    registros: list[dict[str, object]] = []
    for fila in hoja.iter_rows(min_row=encabezado_fila + 1, values_only=True):
        feria = texto(fila[columnas["Feria"]]) if columnas["Feria"] < len(fila) else ""
        if feria.lower().startswith("fuente"):
            break
        fecha = fila[columnas["Fecha"]] if columnas["Fecha"] < len(fila) else None
        if not feria or not isinstance(fecha, (date, datetime)):
            continue
        fecha_real = fecha.date() if isinstance(fecha, datetime) else fecha
        if fecha_real < SERIE_DESDE:
            continue
        fecha_iso = fecha_real.isoformat()
        comuna = texto(fila[columnas["Comuna"]]) if "Comuna" in columnas else ""
        for categoria in CATEGORIAS:
            posicion = columnas.get(categoria)
            if posicion is None or posicion >= len(fila):
                continue
            valor = numero(fila[posicion])
            if valor is not None:
                registros.append(
                    {
                        "categoria": categoria,
                        "feria": feria,
                        "comuna": comuna,
                        "fecha": fecha_iso,
                        "valor": valor,
                        "url": url,
                    }
                )
    return registros


def promedio_por_fecha(registros: list[dict[str, object]]) -> list[dict[str, object]]:
    agrupado: dict[str, list[float]] = defaultdict(list)
    for registro in registros:
        agrupado[str(registro["fecha"])].append(float(registro["valor"]))
    return [
        {"fecha": fecha, "valor": round(mean(valores)), "ferias": len(valores)}
        for fecha, valores in sorted(agrupado.items())
    ]


def historia_operador(registros: list[dict[str, object]], operador: str) -> list[dict[str, object]]:
    return promedio_por_fecha(
        [registro for registro in registros if operador in texto(registro["feria"]).lower()]
    )


def ultimo(historial: list[dict[str, object]]) -> dict[str, object] | None:
    return historial[-1] if historial else None


def cargar_actual() -> dict[str, object]:
    if not DESTINO.exists():
        return {}
    try:
        contenido = json.loads(DESTINO.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return contenido if isinstance(contenido, dict) else {}


def fusionar_historial(anterior: object, nuevo: list[dict[str, object]]) -> list[dict[str, object]]:
    """Conserva la historia y reemplaza sólo las jornadas recién publicadas.

    El dato nacional guardado no se vuelve a usar como si fuese otra feria: eso
    duplicaría una jornada al llegar un boletín nuevo para la misma fecha.
    """
    por_fecha: dict[str, dict[str, object]] = {}
    for fila in anterior if isinstance(anterior, list) else []:
        if not isinstance(fila, dict):
            continue
        try:
            fecha = date.fromisoformat(str(fila.get("fecha")))
            valor = float(fila.get("valor"))
        except (TypeError, ValueError):
            continue
        if fecha >= SERIE_DESDE and valor > 0:
            por_fecha[fecha.isoformat()] = {
                "fecha": fecha.isoformat(),
                "valor": round(valor),
                "ferias": int(fila.get("ferias") or 0),
            }
    for fila in nuevo:
        por_fecha[str(fila["fecha"])] = fila
    return [por_fecha[fecha] for fecha in sorted(por_fecha)]


def construir(historial_completo: bool) -> dict[str, object]:
    actual = cargar_actual()
    registros_por_clave: dict[tuple[str, str, str, str], dict[str, object]] = {}

    fuentes: list[str] = []
    errores: list[str] = []
    for url in links_xlsx(historial_completo):
        try:
            for registro in leer_boletin(descargar(url), url):
                clave = (
                    str(registro["categoria"]),
                    str(registro["feria"]),
                    str(registro["comuna"]),
                    str(registro["fecha"]),
                )
                registros_por_clave[clave] = registro
            fuentes.append(url)
        except Exception as exc:  # Conserva el último JSON válido si un XLSX puntual falla.
            errores.append(f"{url}: {exc}")

    registros = list(registros_por_clave.values())
    categorias_actuales = actual.get("categorias", {})
    if not registros and not isinstance(categorias_actuales, dict):
        raise RuntimeError("No fue posible extraer ni preservar datos oficiales de ODEPA/AFECH.")

    # Conservamos bloques que este lector no gestiona. Así una ejecución
    # programada de Novillo Gordo/Engorda no elimina historia pública de
    # otras categorías que pudiera estar publicada en el mismo JSON.
    categorias: dict[str, object] = (
        dict(categorias_actuales) if isinstance(categorias_actuales, dict) else {}
    )
    for categoria in CATEGORIAS:
        de_categoria = [registro for registro in registros if registro["categoria"] == categoria]
        bloque_anterior = (
            categorias_actuales.get(categoria, {}) if isinstance(categorias_actuales, dict) else {}
        )
        if not isinstance(bloque_anterior, dict):
            bloque_anterior = {}
        nacional = fusionar_historial(
            bloque_anterior.get("nacional"), promedio_por_fecha(de_categoria)
        )
        tattersall = fusionar_historial(
            bloque_anterior.get("tattersall"), historia_operador(de_categoria, "tattersall")
        )
        fegosa = fusionar_historial(
            bloque_anterior.get("fegosa"), historia_operador(de_categoria, "fegosa")
        )
        categorias[categoria] = {
            "nacional": nacional,
            "tattersall": tattersall,
            "fegosa": fegosa,
            "ultimo_nacional": ultimo(nacional),
            "ultimo_tattersall": ultimo(tattersall),
            "ultimo_fegosa": ultimo(fegosa),
        }

    fechas = [
        str(fila["fecha"])
        for categoria in CATEGORIAS
        for fila in categorias[categoria]["nacional"]  # type: ignore[index]
    ]
    if not fechas:
        raise RuntimeError("No hay historia pública suficiente para las categorías configuradas.")
    ultima_fecha = max(fechas)
    return {
        "fuente": "ODEPA / Asociación Gremial de Ferias Ganaderas de Chile (AFECH)",
        "metodo": (
            "Precio promedio por kilo de los 5 primeros precios según feria, nominal "
            "sin IVA. KleoGuard calcula un promedio simple de las ferias por jornada."
        ),
        "serie_desde": SERIE_DESDE.isoformat(),
        "ultima_jornada_publicada": ultima_fecha,
        "boletines_consultados": fuentes,
        "advertencias": errores,
        "categorias": categorias,
    }


def comparable(contenido: dict[str, object]) -> str:
    """Representación estable de los datos que efectivamente se publican.

    Los enlaces consultados y advertencias técnicas pueden variar aunque los
    precios sean idénticos (por ejemplo, por una caída temporal de ODEPA). No
    deben provocar un commit ni una nueva publicación por sí solos.
    """
    relevante = {
        clave: contenido.get(clave)
        for clave in (
            "fuente",
            "metodo",
            "serie_desde",
            "ultima_jornada_publicada",
            "categorias",
        )
    }
    return json.dumps(relevante, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def guardar_si_cambio(resultado: dict[str, object]) -> bool:
    actual = cargar_actual()
    actual_comparable = {
        clave: valor
        for clave, valor in actual.items()
        if clave != "actualizado_en_utc"
    }
    if comparable(actual_comparable) == comparable(resultado):
        return False
    resultado["actualizado_en_utc"] = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    DESTINO.write_text(json.dumps(resultado, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return True


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--historial-completo",
        action="store_true",
        help="Revisa todos los XLSX disponibles en la página pública; úsalo sólo para una reconstrucción.",
    )
    args = parser.parse_args()
    resultado = construir(args.historial_completo)
    cambio = guardar_si_cambio(resultado)
    estado = "actualizado" if cambio else "sin cambios oficiales"
    print(
        f"Mercado {estado}: {resultado['ultima_jornada_publicada']} · "
        f"{len(resultado['boletines_consultados'])} boletines consultados"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
